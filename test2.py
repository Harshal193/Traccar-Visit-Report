import numpy as np
import pandas as pd
import xlrd
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font , Color, Alignment, colors

data_file = pd.read_excel('test_workbookmd.xlsx')
#print(data_file)
pf = pd.DataFrame(data_file, columns=['servertime','type','name','name.1'])
#print(pf)
df = data_file.iloc[1:6]#columns
#print(df)
dn = data_file['type']
#print(dn)
i=1

wb = Workbook()
ws = wb.active
#toro = pf.count(,None,False)
#toro =len(pf.columns)
m = pf.shape[0]               # to display total no of rows in file
print(m)
#dq = df.head
#print(df)

if(data_file.iloc[i, 1] == 'geofenceEnter'):
    for i in range(1,m-1,2):
        j=i+1
        #for j in range(35):
        da = data_file.iloc[i,0]
        db = data_file.iloc[j,0]
        #print(dz,dx,j)
        a = db-da
        b = a.seconds/60
        data = []
        sata = []
        #print(b)
        val = [i]
        if(b > 1):
            #print("Time greater than 5 minutes")
            
            #print(i)
            #data.append(data_file.iloc[i])
            
            #ws.append(data_file.iloc[i])
            #for row in data:
             #   ws.append(row)
            #ws.append(data_file.iloc[i, 0])
        #print(val)
            res = list(data_file.iloc[i,:])
            #res = list(data_file.iloc[i, [0,2,3]])

            #ws.date_range('2019-10-22','2019-11-30',freq='H')
            
            bold_font = Font(bold=True)
            red_text = Font(color=colors.RED, size=20)
            SUBTITLE_ALIGN = Alignment(horizontal='center')

            ws['A2'].font = red_text
            #ws.merge_cells('G1:H2')
            ws['A2'] = 'Visit Report'
            ws.merge_cells('A2:Q2')
            ws['A6'].font = bold_font
            ws['A6'] = 'From:'
            ws['D6'].font = bold_font
            ws['D6'] = 'To:'
            
            ws['A2'].alignment = SUBTITLE_ALIGN
            ws['B10'].alignment = SUBTITLE_ALIGN

            
            ws['A10'].font = bold_font
            ws['B10'].font = bold_font
            ws['C10'].font = bold_font
            ws['D10'].font = bold_font
            ws['A10']='Customer Visit Time'
            ws['B10']='Type'
            ws['C10']='Person'
            ws['D10']='Visited Office'
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 20


            ws.append(res)
            workbook_name = "final_workbook"
            wb.save(workbook_name + ".xlsx")

            print(i)
            print(res)
            print("Data successfully inserted in excel file")

            
            if(data_file.iloc[i, 1] == 'geofenceEnter'):
                print("YES")
            else:
                print("NO")
            #print(val)
else:
    i=i+1
    print("value of i incremented by 1")


            #print(data_file.iloc[i, 0])
        #while(None in data):
            #data.removeall("")
        #for val in data:
            ##if val == None :
                #sata.append(val)
                #sata.remove()
        #print(data)
        #print("workbook created")
            
            
